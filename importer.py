# -*- coding: utf-8 -*-
"""
Import des feuilles Excel "SUIVI DES SINISTRES" vers la base SQLite.
Gère les variantes d'intitulés de colonnes selon les années.
"""
import os
import re
import shutil
import logging
import unicodedata
import datetime
import openpyxl
from openpyxl.styles import PatternFill

import database as db

logger = logging.getLogger(__name__)

# Mots-clés (normalisés) -> nom de colonne interne.
# On associe chaque variante d'intitulé rencontrée dans le fichier au champ normalisé.
HEADER_MAP = {
    "n": "numero",
    "nordre": "numero",
    "chauffeur": "chauffeur",
    "ncodedecam": "code_cam",
    "typedevehicul": "type_vehicule",
    "datedusinistre": "date_sinistre",
    "datedeladeclaration": "date_declaration",
    "lieudaccident": "lieu_accident",
    "ndimatriculation": "immatriculation",
    "nomprenomdechauff": "chauffeur",
    "nomprenomduchauffeur": "chauffeur",
    "typedaccident": "type_accident",
    "degatscause": "degats_cause",
    "avecousanstiers": "avec_sans_tiers",
    "fautifoupasfautif": "fautif",
    "visadereparation": "visa_reparation",
    "expertise": "expertise",
    "dateexpertise": "date_expertise",
    "datedexpertise": "date_expertise",
    "datexpertise": "date_expertise",
    "dateexpert": "date_expertise",
    "pvrecu": "pv_recu",
    "datedereceptiondupv": "date_reception_pv",
    "datedereception": "date_reception_pv",
    "delaisdespvparjour": "delai_pv_jours",
    "confirmationdespv": "confirmation_pv",
    "datedeconfirmationdespv": "date_confirmation_pv",
    "ndossier": "numero_dossier",
    "montantaremboursserselonlepvexpert": "montant_pv_expert",
    "montantaremboursserselonlesachats": "montant_achats",
    "montantdureglementavantlapplicationdelarp": "montant_reglement_avant_rp",
    "montantdureglement": "montant_reglement_avant_rp",
    "ecart": "ecart",
    "banque": "banque",
    "bankagb": "banque",
    "ndecheque": "numero_cheque",
    "datedereglerment": "date_reglement",
    "datedereglement": "date_reglement",
    "datedureglement": "date_reglement",
    "nbrdejoursdesimmobilisation": "jours_immobilisation",
    "nombredheursmaindoeuvr": "heures_maindoeuvre",
    "nombredheurs": "heures_maindoeuvre",
    "montantdepeintur": "montant_peinture",
    "montantdesfourniture": "montant_fournitures",
    "vetuste": "vetuste",
    "franshise": "franchise",
    "reglerment": "statut_reglement",
    "reglement": "statut_reglement",
    "statut": "statut_reglement",
    "statutdureglement": "statut_reglement",
    "circonstancedaccident": "circonstance_accident",
    "circonstanceaccident": "circonstance_accident",
    "observation": "observation",
    "delaidereg": "delai_reg",
    "ndossiergmax": "numero_dossier_gmax",
    # Champs optionnels (cahier des charges §14/§15) : ne seront renseignés que si
    # une colonne correspondante existe réellement dans le fichier Excel.
    "compagnie": "compagnie",
    "compagniedassurance": "compagnie",
    "agence": "agence",
    "expert": "expert",
    "nomdelexpert": "expert",
    "camion": "camion",
    "nomducamion": "camion",
    "assure": "assure",
    "assurer": "assure",
}

REQUIRED_HEADER_HINTS = ["datedusinistre", "lieudaccident", "nomprenomdechauff", "nomprenomduchauffeur"]


def normalize(text):
    if text is None:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]", "", text)
    return text


def find_header_row(ws, max_scan=10):
    """Cherche la ligne d'en-tête en scannant les premières lignes."""
    for r in range(1, max_scan + 1):
        row_vals = [normalize(c.value) for c in ws[r]]
        hits = sum(1 for hint in REQUIRED_HEADER_HINTS if hint in row_vals)
        if hits >= 1:
            return r
    return None


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, str):
        v = value.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(v, fmt).date().isoformat()
            except ValueError:
                continue
        return None
    return None


def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        v = value.strip().replace(" ", "").replace(",", ".")
        try:
            return float(v)
        except ValueError:
            return None
    return None


def extract_year_from_sheet_name(name):
    m = re.search(r"(19|20)\d{2}", name)
    if m:
        return int(m.group(0))
    return None


def is_cell_in_merged_range(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return True
    return False


def should_preserve_existing_cell(ws, row, col, value):
    if is_cell_in_merged_range(ws, row, col):
        return True
    target = ws.cell(row=row, column=col)
    if target.value is None:
        return False
    if isinstance(target.value, str) and target.value.startswith("="):
        return True
    if isinstance(target.value, (int, float)) and not isinstance(value, (int, float)):
        return False
    return False


def coerce_value_for_field(field, value):
    if field in ("date_sinistre", "date_declaration", "date_expertise", "date_reception_pv", "date_confirmation_pv", "date_reglement"):
        return parse_date(value)
    if field in ("delai_pv_jours", "montant_pv_expert", "montant_reglement_avant_rp", "ecart", "jours_immobilisation", "heures_maindoeuvre", "montant_peinture", "montant_fournitures", "vetuste", "franchise", "delai_reg", "montant_achats"):
        return parse_number(value)
    if isinstance(value, str):
        return value.strip()
    return value


def normalize_for_matching(field, value):
    normalized = coerce_value_for_field(field, value)
    if normalized is None:
        return None
    if isinstance(normalized, datetime.datetime):
        return normalized.date().isoformat()
    if isinstance(normalized, datetime.date):
        return normalized.isoformat()
    if isinstance(normalized, (int, float)):
        return normalized
    return str(normalized).strip()


def find_matching_row(ws, header_row, col_map, record, used_rows):
    field_to_col = {field: idx + 1 for idx, field in col_map.items()}
    candidate_key_sets = [
        ("date_sinistre", "chauffeur"),
        ("numero_dossier", "chauffeur"),
        ("numero", "chauffeur"),
        ("date_sinistre", "numero_dossier"),
        ("date_sinistre", "numero"),
        ("numero_dossier",),
        ("numero",),
        ("date_sinistre",),
    ]

    for key_fields in candidate_key_sets:
        if not any(record.get(field) not in (None, "") for field in key_fields):
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if row_idx in used_rows:
                continue
            matches = True
            for field in key_fields:
                if field not in field_to_col:
                    continue
                existing_value = ws.cell(row=row_idx, column=field_to_col[field]).value
                if normalize_for_matching(field, existing_value) != normalize_for_matching(field, record.get(field)):
                    matches = False
                    break
            if matches:
                return row_idx

    return None


def find_empty_data_row(ws, header_row, col_map, used_rows):
    """Cherche une ligne existante totalement vide (hors cellules fusionnées) pour la
    réutiliser avant d'ajouter une nouvelle ligne en fin de feuille. Cela évite de créer
    des lignes superflues quand la feuille contient déjà des lignes vides ou fusionnées."""
    field_to_col = {field: idx + 1 for idx, field in col_map.items()}
    cols = list(field_to_col.values())
    if not cols:
        return None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if row_idx in used_rows:
            continue
        all_empty = True
        for col in cols:
            if is_cell_in_merged_range(ws, row_idx, col):
                continue
            cell_value = ws.cell(row=row_idx, column=col).value
            if cell_value not in (None, ""):
                all_empty = False
                break
        if all_empty:
            return row_idx
    return None


def normalize_order_value(value):
    if value is None:
        return (1, "")
    text = str(value).strip()
    if not text:
        return (1, "")
    m = re.search(r"(\d+)", text)
    if m:
        return (0, int(m.group(1)))
    return (1, text)


def find_row_by_numero_placeholder(ws, header_row, col_map, record, used_rows):
    """Cherche une ligne déjà pré-numérotée (colonne N°) correspondant exactement
    au numéro du nouveau sinistre, même si le reste de la ligne est vide.
    Cas fréquent : le fichier Excel a des lignes pré-numérotées à l'avance
    (ex. N°90, N°91... déjà écrits en colonne A alors que le sinistre n'a pas
    encore eu lieu). Sans cette recherche, ces lignes semblent "non vides" et
    le nouveau sinistre serait ajouté beaucoup plus loin dans la feuille."""
    numero_col = None
    for idx, field in col_map.items():
        if field == "numero":
            numero_col = idx + 1
            break
    if numero_col is None:
        return None
    target = normalize_order_value(record.get("numero"))
    if target[0] != 0:  # pas de numéro exploitable (texte libre ou vide)
        return None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if row_idx in used_rows:
            continue
        existing_value = ws.cell(row=row_idx, column=numero_col).value
        if normalize_order_value(existing_value) == target:
            return row_idx
    return None


def sync_records_to_workbook(filepath, records, progress_callback=None):
    """Synchronise les données vers le fichier Excel source sans modifier sa structure,
    ses formules ni ses formats. Retourne (backup_path, assignments, unmatched) où
    assignments est une liste de (record_id, sheet_name, row_idx) pour les lignes
    créées/retrouvées, et unmatched la liste des sinistres qui n'ont pu être placés
    dans aucune feuille (année manquante ou classeur sans feuille reconnue)."""
    if not filepath or not os.path.exists(filepath):
        raise FileNotFoundError("Fichier Excel source introuvable")

    backup_dir = os.path.join(os.path.dirname(filepath), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f"{os.path.splitext(os.path.basename(filepath))[0]}_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(filepath)[1]}")
    shutil.copy2(filepath, backup_path)

    wb = openpyxl.load_workbook(filepath, data_only=False)

    # Années réellement disponibles comme feuilles reconnues dans le classeur
    available_years = []
    for sheet_name in wb.sheetnames:
        ws_probe = wb[sheet_name]
        if find_header_row(ws_probe) is None:
            continue
        y = extract_year_from_sheet_name(sheet_name)
        if y is not None:
            available_years.append(y)

    def resolve_year(record):
        """Retourne l'année à utiliser pour ce sinistre, en la déduisant de la date
        du sinistre si le champ Année est resté vide."""
        year = record.get("annee")
        if year not in (None, ""):
            try:
                return int(year)
            except (TypeError, ValueError):
                pass
        ds = record.get("date_sinistre")
        if ds:
            try:
                return datetime.date.fromisoformat(str(ds)).year
            except ValueError:
                pass
        return None

    records_by_year = {}
    unmatched = []
    for record in records:
        year = resolve_year(record)
        if year is None:
            unmatched.append(record)
            continue
        target_year = year
        if available_years and year not in available_years:
            # Aucune feuille pour cette année précise : on rattache à la feuille de
            # l'année existante la plus proche, plutôt que de perdre le sinistre.
            target_year = min(available_years, key=lambda y: abs(y - year))
        elif not available_years:
            unmatched.append(record)
            continue
        records_by_year.setdefault(target_year, []).append(record)

    for year, year_records in records_by_year.items():
        year_records.sort(key=lambda r: normalize_order_value(r.get("numero")))

    detected_colors = db.load_status_colors()

    def get_row_fill(statut):
        normalized = (statut or "").strip().upper()
        if normalized in detected_colors:
            hexcolor = detected_colors[normalized].lstrip("#")
            return PatternFill(fill_type="solid", fgColor="00" + hexcolor)
        if normalized == "REGLER":
            return PatternFill(fill_type="solid", fgColor="00E8F5E9")
        if normalized in {"EN COURS", "ENCOURS", "EN-COURS", "NON REGLER", "NON REGLÉ", "INSTANCE"}:
            return PatternFill(fill_type="solid", fgColor="00FFF8E1")
        if normalized in {"NEANT", "NÉANT", "AUCUN", ""}:
            return PatternFill(fill_type="solid", fgColor="00F3F4F6")
        return PatternFill(fill_type="solid", fgColor="00F1F3F5")

    assignments = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue

        year = extract_year_from_sheet_name(sheet_name)
        year_records = records_by_year.get(year, [])
        header_cells = ws[header_row]
        col_map = {}
        for idx, cell in enumerate(header_cells):
            key = normalize(cell.value)
            if key in HEADER_MAP:
                col_map[idx] = HEADER_MAP[key]

        used_rows = set()
        for record in year_records:
            target_row_idx = None

            # Chemin rapide : on connaît déjà la ligne exacte de ce sinistre dans cette feuille.
            known_sheet = record.get("excel_sheet")
            known_row = record.get("excel_row")
            if known_sheet == sheet_name and known_row and known_row not in used_rows \
                    and known_row <= ws.max_row and known_row > header_row:
                target_row_idx = known_row

            if target_row_idx is None:
                target_row_idx = find_matching_row(ws, header_row, col_map, record, used_rows)
            if target_row_idx is None:
                target_row_idx = find_row_by_numero_placeholder(ws, header_row, col_map, record, used_rows)
            if target_row_idx is None:
                target_row_idx = find_empty_data_row(ws, header_row, col_map, used_rows)
            if target_row_idx is None:
                target_row_idx = ws.max_row + 1
                ws.append([])
            used_rows.add(target_row_idx)
            assignments.append((record.get("id"), sheet_name, target_row_idx))

            for col_idx, field in col_map.items():
                value = coerce_value_for_field(field, record.get(field))
                target_col = col_idx + 1
                if value in (None, ""):
                    continue
                if should_preserve_existing_cell(ws, target_row_idx, target_col, value):
                    continue

                existing_value = ws.cell(row=target_row_idx, column=target_col).value
                if normalize_for_matching(field, existing_value) == normalize_for_matching(field, value):
                    continue
                ws.cell(row=target_row_idx, column=target_col, value=value)

            statut = (record.get("statut_reglement") or "").strip().upper()
            fill = get_row_fill(statut)
            if ws.max_column > 0:
                for col in range(1, ws.max_column + 1):
                    if is_cell_in_merged_range(ws, target_row_idx, col):
                        continue
                    target = ws.cell(row=target_row_idx, column=col)
                    if target.value is None:
                        target.value = ""
                    if not isinstance(target.value, str) or not target.value.startswith("="):
                        target.fill = fill

        if progress_callback:
            progress_callback(f"🔄 {sheet_name} : {len(year_records)} ligne(s) synchronisée(s)")

    wb.save(filepath)
    return backup_path, assignments, unmatched


def clear_excel_row(filepath, sheet_name, row_idx, progress_callback=None):
    """Efface le contenu d'une ligne précise (suppression définitive depuis la Corbeille).
    Ne touche pas aux cellules fusionnées ni aux formules ; crée une sauvegarde avant."""
    if not filepath or not os.path.exists(filepath):
        raise FileNotFoundError("Fichier Excel source introuvable")

    backup_dir = os.path.join(os.path.dirname(filepath), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f"{os.path.splitext(os.path.basename(filepath))[0]}_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(filepath)[1]}")
    shutil.copy2(filepath, backup_path)

    wb = openpyxl.load_workbook(filepath, data_only=False)
    if sheet_name not in wb.sheetnames:
        return backup_path
    ws = wb[sheet_name]
    if row_idx < 1 or row_idx > ws.max_row:
        return backup_path

    for col in range(1, ws.max_column + 1):
        if is_cell_in_merged_range(ws, row_idx, col):
            continue
        cell = ws.cell(row=row_idx, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            continue
        cell.value = None
        cell.fill = PatternFill(fill_type=None)

    wb.save(filepath)
    if progress_callback:
        progress_callback(f"🗑 Ligne {row_idx} effacée dans {sheet_name}")
    return backup_path


def clear_all_data_workbook(filepath, progress_callback=None):
    """Vide toutes les données des feuilles reconnues, en conservant les feuilles, les
    en-têtes, les formules, la mise en forme, les couleurs et les listes de validation.
    Utilisé par Administration > Vider le fichier Excel. Crée une sauvegarde avant."""
    if not filepath or not os.path.exists(filepath):
        raise FileNotFoundError("Fichier Excel source introuvable")

    backup_dir = os.path.join(os.path.dirname(filepath), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f"{os.path.splitext(os.path.basename(filepath))[0]}_avant_vidage_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(filepath)[1]}")
    shutil.copy2(filepath, backup_path)

    wb = openpyxl.load_workbook(filepath, data_only=False)
    total_cleared = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if is_cell_in_merged_range(ws, row_idx, col):
                    continue
                cell = ws.cell(row=row_idx, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue  # on conserve les formules
                if cell.value not in (None, ""):
                    cell.value = None
                    total_cleared += 1
        if progress_callback:
            progress_callback(f"🧹 {sheet_name} : données effacées (en-têtes et mise en forme conservés)")

    wb.save(filepath)
    return backup_path, total_cleared


def _cell_fill_hex(cell):
    """Retourne la couleur de remplissage d'une cellule sous forme '#RRGGBB', ou None."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type != "solid":
            return None
        color = fill.fgColor
        if color is None:
            return None
        rgb = color.rgb
        if not rgb or not isinstance(rgb, str):
            return None
        rgb = rgb[-6:]
        if rgb.upper() in ("FFFFFF", "000000"):
            return None  # blanc/noir : probablement pas une couleur de statut significative
        return "#" + rgb.upper()
    except Exception:
        # Pas de couleur détectable (cellule sans remplissage solide, couleur
        # malformée...) : best-effort, on retourne None pour retomber sur les
        # couleurs par défaut. Debug uniquement (appelé très souvent).
        logger.debug("Aucune couleur détectable sur la cellule %s", cell.coordinate, exc_info=True)
        return None


def detect_status_colors(ws, header_row, col_map):
    """Échantillonne la couleur de remplissage réelle des lignes du fichier Excel,
    par statut, pour reproduire fidèlement les couleurs dans l'application."""
    statut_col = None
    for idx, field in col_map.items():
        if field == "statut_reglement":
            statut_col = idx + 1
            break
    if statut_col is None:
        return {}

    from collections import Counter
    samples = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        statut_val = ws.cell(row=row_idx, column=statut_col).value
        statut = (str(statut_val).strip().upper() if statut_val else "")
        if not statut:
            continue
        color = _cell_fill_hex(ws.cell(row=row_idx, column=statut_col))
        if color is None:
            continue
        samples.setdefault(statut, Counter())[color] += 1

    result = {}
    for statut, counter in samples.items():
        result[statut] = counter.most_common(1)[0][0]
    return result


def import_workbook(filepath, progress_callback=None, clear_existing=False):
    """
    Importe toutes les feuilles reconnues du classeur.
    progress_callback(message: str) est appelé pour suivre la progression.
    Retourne un résumé {feuille: nb_importés}.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    db.init_db()
    summary = {}
    all_status_colors = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            if progress_callback:
                progress_callback(f"⏭ Feuille ignorée (en-tête introuvable) : {sheet_name}")
            continue

        year = extract_year_from_sheet_name(sheet_name)

        # Construire le mapping colonne_index -> champ_interne
        col_map = {}
        for idx, cell in enumerate(ws[header_row]):
            key = normalize(cell.value)
            if key in HEADER_MAP:
                col_map[idx] = HEADER_MAP[key]

        if not col_map:
            if progress_callback:
                progress_callback(f"⏭ Feuille ignorée (aucune colonne reconnue) : {sheet_name}")
            continue

        if clear_existing and year:
            db.clear_year(year)

        # Détection des couleurs réelles utilisées pour chaque statut dans cette feuille
        sheet_colors = detect_status_colors(ws, header_row, col_map)
        for statut, color in sheet_colors.items():
            all_status_colors.setdefault(statut, color)

        records = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if row is None or all(v is None or v == "" for v in row):
                continue
            record = {"annee": year, "source_sheet": sheet_name, "excel_sheet": sheet_name, "excel_row": row_idx}
            has_key_data = False
            for idx, field in col_map.items():
                if idx >= len(row):
                    continue
                value = row[idx]
                if field in ("date_sinistre", "date_declaration", "date_expertise",
                             "date_reception_pv", "date_confirmation_pv", "date_reglement"):
                    value = parse_date(value)
                elif field in ("delai_pv_jours", "montant_pv_expert", "montant_reglement_avant_rp",
                               "ecart", "jours_immobilisation", "heures_maindoeuvre",
                               "montant_peinture", "montant_fournitures", "vetuste",
                               "franchise", "delai_reg", "montant_achats"):
                    value = parse_number(value)
                elif isinstance(value, str):
                    value = value.strip()
                if field == "chauffeur" and value:
                    has_key_data = True
                if field == "date_sinistre" and value:
                    has_key_data = True
                record[field] = value
            if has_key_data:
                records.append(record)

        inserted = db.bulk_insert(records)
        summary[sheet_name] = inserted
        if progress_callback:
            progress_callback(f"✔ {sheet_name} : {inserted} sinistre(s) importé(s) sur {len(records)} lu(s)")

    if all_status_colors:
        db.save_status_colors(all_status_colors)
        if progress_callback:
            progress_callback(f"🎨 Couleurs détectées pour {len(all_status_colors)} statut(s) : "
                               f"{', '.join(all_status_colors.keys())}")

    return summary
