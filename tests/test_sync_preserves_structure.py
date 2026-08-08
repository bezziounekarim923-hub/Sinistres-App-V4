import os
import shutil
import tempfile
import unittest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import importer


class SyncPreservesStructureTest(unittest.TestCase):
    def test_sync_keeps_existing_rows_and_styles(self):
        temp_dir = tempfile.mkdtemp(prefix="sinistres_test_", dir=os.getcwd())
        self.addCleanup(shutil.rmtree, temp_dir, ignore_errors=True)
        path = os.path.join(temp_dir, "source.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "SUIVI DES SINISTRES 2024"
        ws.append(["Date", "Chauffeur", "Statut"])
        ws.cell(row=2, column=1, value="2024-01-01")
        ws.cell(row=2, column=2, value="Alice")
        ws.cell(row=2, column=3, value="EN COURS")
        ws.cell(row=3, column=1, value="2024-02-01")
        ws.cell(row=3, column=2, value="Bob")
        ws.cell(row=3, column=3, value="REGLER")

        ws.cell(row=1, column=1).font = Font(bold=True, color="FF0000")
        ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        ws.row_dimensions[2].height = 25
        wb.save(path)

        importer.sync_records_to_workbook(
            path,
            [{"annee": 2024, "date_sinistre": "2024-01-01", "chauffeur": "Alice", "statut_reglement": "REGLER"},
             {"annee": 2024, "date_sinistre": "2024-02-01", "chauffeur": "Bob", "statut_reglement": "EN COURS"}],
        )

        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        ws2 = wb2["SUIVI DES SINISTRES 2024"]
        self.assertEqual(ws2.max_row, 3)
        self.assertEqual(ws2.row_dimensions[2].height, 25)
        self.assertEqual(ws2.cell(row=1, column=1).font.bold, True)
        self.assertEqual(ws2.cell(row=1, column=1).fill.fgColor.rgb, "00FFFF00")

    def test_sync_orders_rows_by_order_number(self):
        temp_dir = tempfile.mkdtemp(prefix="sinistres_test_", dir=os.getcwd())
        self.addCleanup(shutil.rmtree, temp_dir, ignore_errors=True)
        path = os.path.join(temp_dir, "source_order.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "SUIVI DES SINISTRES 2024"
        ws.append(["Date du sinistre", "Lieu d'accident", "Nordre", "Chauffeur"])
        wb.save(path)

        importer.sync_records_to_workbook(
            path,
            [
                {"annee": 2024, "numero": "N 90", "chauffeur": "Zulu"},
                {"annee": 2024, "numero": "N 01", "chauffeur": "Alpha"},
                {"annee": 2024, "numero": "N 03", "chauffeur": "Charlie"},
            ],
        )

        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        ws2 = wb2["SUIVI DES SINISTRES 2024"]
        self.assertEqual(ws2.cell(row=2, column=3).value, "N 01")
        self.assertEqual(ws2.cell(row=3, column=3).value, "N 03")
        self.assertEqual(ws2.cell(row=4, column=3).value, "N 90")

    def test_sync_ignores_merged_cells(self):
        temp_dir = tempfile.mkdtemp(prefix="sinistres_test_", dir=os.getcwd())
        self.addCleanup(shutil.rmtree, temp_dir, ignore_errors=True)
        path = os.path.join(temp_dir, "source_merged.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "SUIVI DES SINISTRES 2024"
        ws.append(["Date du sinistre", "Chauffeur", "Statut"])
        ws.merge_cells("A2:A3")
        wb.save(path)

        importer.sync_records_to_workbook(
            path,
            [{"annee": 2024, "date_sinistre": "2024-01-01", "chauffeur": "Alice", "statut_reglement": "REGLER"}],
        )

        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        ws2 = wb2["SUIVI DES SINISTRES 2024"]
        self.assertEqual(ws2.max_row, 3)

    def test_sync_preserves_formula_cells(self):
        temp_dir = tempfile.mkdtemp(prefix="sinistres_test_", dir=os.getcwd())
        self.addCleanup(shutil.rmtree, temp_dir, ignore_errors=True)
        path = os.path.join(temp_dir, "source_formula.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "SUIVI DES SINISTRES 2024"
        ws.append(["Date du sinistre", "Chauffeur", "Statut"])
        ws.cell(row=3, column=3, value="=1+1")
        wb.save(path)

        importer.sync_records_to_workbook(
            path,
            [{"annee": 2024, "date_sinistre": "2024-01-01", "chauffeur": "Alice", "statut_reglement": "REGLER"}],
        )

        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        ws2 = wb2["SUIVI DES SINISTRES 2024"]
        self.assertEqual(ws2.cell(row=3, column=3).value, "=1+1")

    def test_sync_updates_only_matching_row(self):
        temp_dir = tempfile.mkdtemp(prefix="sinistres_test_", dir=os.getcwd())
        self.addCleanup(shutil.rmtree, temp_dir, ignore_errors=True)
        path = os.path.join(temp_dir, "source_targeted.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "SUIVI DES SINISTRES 2024"
        ws.append(["Date du sinistre", "Chauffeur", "Statut"])
        ws.cell(row=2, column=1, value="2024-01-01")
        ws.cell(row=2, column=2, value="Alice")
        ws.cell(row=2, column=3, value="EN COURS")
        ws.cell(row=3, column=1, value="2024-02-01")
        ws.cell(row=3, column=2, value="Bob")
        ws.cell(row=3, column=3, value="EN COURS")
        wb.save(path)

        importer.sync_records_to_workbook(
            path,
            [{"annee": 2024, "date_sinistre": "2024-01-01", "chauffeur": "Alice", "statut_reglement": "REGLER"}],
        )

        from openpyxl import load_workbook
        wb2 = load_workbook(path)
        ws2 = wb2["SUIVI DES SINISTRES 2024"]
        self.assertEqual(ws2.cell(row=2, column=3).value, "REGLER")
        self.assertEqual(ws2.cell(row=3, column=3).value, "EN COURS")


if __name__ == "__main__":
    unittest.main()
